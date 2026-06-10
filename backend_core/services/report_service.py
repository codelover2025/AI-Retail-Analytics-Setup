"""
Report generation service — Module 3 (Phase 4).

Generates PDF, Excel, and CSV reports from dashboard aggregation data.
Runs synchronously in a thread pool — called by the scheduler or directly.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

from sqlalchemy.orm import Session

from backend_core.services.dashboard_service import DashboardService
from shared.config import Settings
from shared.database.report_models import ReportJob

logger = logging.getLogger(__name__)


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


class ReportService:
    """Generates formatted reports from analytics data."""

    def __init__(self, db: Session, settings: Settings, brand_id: uuid.UUID) -> None:
        self.db = db
        self.settings = settings
        self.brand_id = brand_id
        self._output_dir = Path(settings.reports_output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Job lifecycle
    # ------------------------------------------------------------------

    def create_job(
        self,
        report_type: str,
        output_format: str,
        store_ids: list[str],
        params: dict[str, Any],
        requested_by: Optional[str] = None,
    ) -> ReportJob:
        job = ReportJob(
            brand_id=self.brand_id,
            store_ids=store_ids,
            report_type=report_type,
            output_format=output_format,
            status="pending",
            params=params,
            requested_by=requested_by,
        )
        self.db.add(job)
        self.db.commit()
        self.db.refresh(job)
        return job

    def get_job(self, job_id: uuid.UUID) -> Optional[ReportJob]:
        return self.db.get(ReportJob, job_id)

    # ------------------------------------------------------------------
    # Data collection
    # ------------------------------------------------------------------

    def _collect_data(
        self,
        store_ids: list[str],
        from_day: date,
        to_day: date,
    ) -> dict[str, Any]:
        svc = DashboardService(self.db, self.settings, self.brand_id)
        return svc.overview(
            from_day=from_day,
            to_day=to_day,
            store_ids=store_ids or None,
        )

    def _date_range(self, report_type: str, params: dict) -> tuple[date, date]:
        today = date.today()
        if report_type == "daily":
            d = date.fromisoformat(params.get("date", str(today)))
            return d, d
        elif report_type == "weekly":
            to_d = today
            from_d = today - timedelta(days=6)
            return from_d, to_d
        elif report_type == "monthly":
            to_d = today
            from_d = today.replace(day=1)
            return from_d, to_d
        else:  # custom
            from_d = date.fromisoformat(params.get("from_day", str(today - timedelta(days=29))))
            to_d = date.fromisoformat(params.get("to_day", str(today)))
            return from_d, to_d

    # ------------------------------------------------------------------
    # CSV generation
    # ------------------------------------------------------------------

    def _generate_csv(self, data: dict[str, Any]) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Store ID", "Total Visitors", "Repeat Visitors", "New Visitors",
                         "Repeat Ratio", "Avg Dwell (s)", "Staff Interactions", "Top Zone"])
        for store in data.get("stores", []):
            writer.writerow([
                store.get("store_id"),
                store.get("total_visitors"),
                store.get("repeat_visitors"),
                store.get("new_visitors"),
                store.get("repeat_ratio"),
                store.get("avg_dwell_seconds"),
                store.get("staff_interactions"),
                store.get("top_zone", ""),
            ])
        return buf.getvalue().encode("utf-8")

    # ------------------------------------------------------------------
    # Excel generation
    # ------------------------------------------------------------------

    def _generate_excel(self, data: dict[str, Any], report_type: str) -> bytes:
        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
        except ImportError:
            logger.error("openpyxl not installed — cannot generate Excel report")
            raise

        wb = Workbook()
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Orzen Vision — Retail Analytics Report"])
        ws_summary.append([f"Report Type: {report_type.title()}"])
        ws_summary.append([f"From: {data.get('from_day')}   To: {data.get('to_day')}"])
        ws_summary.append([])
        summary = data.get("summary", {})
        for k, v in summary.items():
            ws_summary.append([k.replace("_", " ").title(), v])
        ws_summary.append([])

        # Store detail sheet
        ws_stores = wb.create_sheet("Stores")
        headers = ["Store ID", "Total Visitors", "Repeat Visitors", "New Visitors",
                   "Repeat Ratio", "Avg Dwell (s)", "Staff Interactions", "Top Zone"]
        ws_stores.append(headers)
        header_font = Font(bold=True)
        for cell in ws_stores[1]:
            cell.font = header_font

        for store in data.get("stores", []):
            ws_stores.append([
                store.get("store_id"),
                store.get("total_visitors"),
                store.get("repeat_visitors"),
                store.get("new_visitors"),
                store.get("repeat_ratio"),
                store.get("avg_dwell_seconds"),
                store.get("staff_interactions"),
                store.get("top_zone", ""),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # PDF generation
    # ------------------------------------------------------------------

    def _generate_pdf(self, data: dict[str, Any], report_type: str) -> bytes:
        try:
            from reportlab.lib import colors  # type: ignore
            from reportlab.lib.pagesizes import A4  # type: ignore
            from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
            from reportlab.lib.units import mm  # type: ignore
            from reportlab.platypus import (  # type: ignore
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
        except ImportError:
            logger.error("reportlab not installed — cannot generate PDF report")
            raise

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        story.append(Paragraph("Orzen Vision — Retail Analytics Report", styles["Title"]))
        story.append(Paragraph(f"Report Type: {report_type.title()}", styles["Normal"]))
        story.append(Paragraph(
            f"Period: {data.get('from_day')} to {data.get('to_day')}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 6 * mm))

        # Summary table
        summary = data.get("summary", {})
        if summary:
            story.append(Paragraph("Brand Summary", styles["Heading2"]))
            summary_data = [["Metric", "Value"]] + [
                [k.replace("_", " ").title(), str(v)] for k, v in summary.items()
            ]
            t = Table(summary_data, colWidths=[80 * mm, 60 * mm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f0f0")]),
            ]))
            story.append(t)
            story.append(Spacer(1, 6 * mm))

        # Stores table
        stores = data.get("stores", [])
        if stores:
            story.append(Paragraph("Store Performance", styles["Heading2"]))
            headers = ["Store", "Visitors", "Repeat", "Dwell(s)", "Interactions"]
            rows = [headers] + [
                [
                    s.get("store_id", ""),
                    str(s.get("total_visitors", 0)),
                    str(s.get("repeat_visitors", 0)),
                    str(round(s.get("avg_dwell_seconds", 0), 1)),
                    str(s.get("staff_interactions", 0)),
                ]
                for s in stores
            ]
            t2 = Table(rows)
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16213e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ]))
            story.append(t2)

        doc.build(story)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Main execute entry point
    # ------------------------------------------------------------------

    def execute_job(self, job_id: uuid.UUID) -> str:
        """Run a report job synchronously; returns file path on success."""
        job = self.get_job(job_id)
        if job is None:
            raise ValueError(f"Job {job_id} not found")

        job.status = "running"
        job.started_at = _utcnow()
        self.db.commit()

        try:
            from_day, to_day = self._date_range(job.report_type, job.params or {})
            data = self._collect_data(list(job.store_ids or []), from_day, to_day)

            if job.output_format == "csv":
                content = self._generate_csv(data)
                ext = "csv"
            elif job.output_format == "excel":
                content = self._generate_excel(data, job.report_type)
                ext = "xlsx"
            else:
                content = self._generate_pdf(data, job.report_type)
                ext = "pdf"

            filename = f"{job.id}_{job.report_type}.{ext}"
            file_path = str(self._output_dir / filename)
            with open(file_path, "wb") as f:
                f.write(content)

            job.status = "completed"
            job.file_path = file_path
            job.completed_at = _utcnow()
            self.db.commit()
            return file_path

        except Exception as exc:
            logger.error("Report job %s failed: %s", job_id, exc)
            job.status = "failed"
            job.error_message = str(exc)
            job.completed_at = _utcnow()
            self.db.commit()
            raise
